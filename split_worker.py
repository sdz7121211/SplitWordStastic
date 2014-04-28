# -*- encoding: utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding("utf-8")

from split_util import split_worker
from openpyxl import Workbook
from pymongo import MongoClient

# mongo_con = MongoClient("103.29.133.167", 27017)

mongo_con = MongoClient("103.29.133.171", 30001)

tasks = {
        "coolpad": {
            "db": "bbscrawl1",
            "collection": "attention",
            "find_": {"forum_id": {"$in": ["xiaolajiaoqm3", "xiaolajiao3"]}},
            "col_name": "attention"},

        "nubiya": {
            "db": "bbscrawl1",
            "collection": "comments",
            "find_": {"forum_id": "nubiya"},
            "col_name": "comment"},

        "jiwu": {
            "db": "bbscrawl1",
            "collection": "comments",
            "find_": {"forum_id": "jiwu"},
            "col_name": "comment"},
        
        "xiaolajiao": {
            "db": "bbscrawl1",
            "collection": "profile",
            "find_": {"forum_id": "xiaolajiao", "sex": {"$exists": True}},
            "col_name": "sex"},

        "thl": {
            "db": "bbscrawl1",
            "collection": "comments",
            "find_": {"forum_id": "thl"},
            "col_name": "comment"},

        "zhidao_xiaomi": {
            "db": "zhidaocrawl",
            "collection": "comments",
            "find_": {"forum_id": "xiaomi"},
            "col_name": "comment"},

        "collect_machinetype": {
            "db": "bbscrawl1",
            "collection": "profile",
            "find_": {"website": "xiaolajiaoshequ"},
            "col_name": "machine_type"},

        "collect_sex": {
            "db": "bbscrawl1",
            "collection": "profile",
            "find_": {"website": "xiaolajiaoshequ"},
            "col_name": "sex"},

        "collect_comments": {
            "db": "zhidaocrawl",
            "collection": "comments",
            "find_": {"forum_id": "nubia"},
            "col_name": "comment"}
        }


def iterColumn(task):
    mongo_db = mongo_con[task["db"]]
    mongo_collection_comment = mongo_db[task["collection"]]
    col_name = task.get("col_name")
    count = 5000
    for comment in mongo_collection_comment.find(task["find_"]):
        # if "context" in comment:
        #    comment = comment.split("F")[0]
        if count:
            count = count - 1
            ok = comment.get(col_name)
            if ok:
                yield ok


def saveComments(dic_split):
    wb = Workbook()
    ws = wb.create_sheet(0)
    col_ = 0
    row_ = 0
    for item in dic_split:
        col_ = 0
        word = item
        print word
        ws.cell(row=row_, column=col_).value = word
        row_ = row_ + 1
    print "-----------------------------end----------------------"
    print " ws.cell(0, 0).value", ws.cell(row=0, column=0).value
    wb.save("col_out.xlsx")


def collectItem(canIter):
    wb = Workbook()
    ws = wb.create_sheet(0)
    col_ = 0
    row_ = 0
    result = {}
    for item in canIter:
        if not item:
            continue
        if type(item) == type([]):
            for ite in item:
                ite = ite.strip()
                if "-" in ite:
                    ite = ite.split("-")[0]
                if result.get(ite):
                    result[ite] = result[ite] + 1
                else:
                    result[ite] = 1
        else:
            item = item.split()[0]
            if "-" in item:
                item = item.split("-")[0]
            if result.get(item):
                result[item] = result[item] + 1
            else:
                result[item] = 1
    l = sorted(result.items(), key=lambda result: result[1], reverse=True)
    print l
    col_ = 0
    row_ = 0
    for key in l:
        print key
        col_ = 0
        ws.cell(row=row_, column=col_).value = key[0]
        col_ = col_ + 1
        ws.cell(row=row_, column=col_).value = key[1]
        row_ = row_ + 1
    wb.save("collect_out.xlsx")

if __name__ == "__main__":
    collectItem(iterColumn(tasks["coolpad"]))
    # saveComments(iterColumn(tasks["coolpad"]))
    # split_worker(iterColumn(tasks["coolpad"]))
