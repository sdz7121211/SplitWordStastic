# -*- encoding: utf-8 -*-
from pymongo import MongoClient
import pymongo
from openpyxl import Workbook
# from openpyxl.writer.excel import ExcelWriter
import sys
reload(sys)
sys.setdefaultencoding("utf-8")

mongo_host = "103.29.133.171"
mongo_port = 30001

con = MongoClient(mongo_host, mongo_port)
mongo_con = con

number = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
letter = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "g", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]

# no_contain = top_comment_toexcel_task["xiaolajiao"]["contain"] + top_comment_toexcel_task["honglajiao"]["contain"] + top_comment_toexcel_task["xiaolajiao3"]["contain"]

top_comment_toexcel_task = {
        "xiaolajiao": {
                "db": "bbscrawl1",
                "collection": "post",
                "contain": ["beidouxiaolajiao",\
                "beidouxingxiaolajiao", "xiaolajiao", "xiaolajiaodianxin",\
                "xiaolajiaoi2", "xiaolajiaom1", "xiaolajiaom1y",\
                "xiaolajiaom1s", "xiaolajiaom2", "xiaolajiaom3",\
                "xiaolajiaoq1", "xiaolajiaozhinengshouji", "xiaolajiaoshoujiguanwang",\
                "yuyinxiaolajiao", "yuxin"],
                "output_file": "xiaolajiao_top_comment.xlsx"},
        "honglajiao": {
                "db": "bbscrawl1",
                "collection": "post",
                "contain": ["honglajiaoshouji"],
                "output_file": "honglajiao_top_comment.xlsx"},
        "xiaolajiao3": {
                "db": "bbscrawl1",
                "collection": "post",
                "contain": ["xiaolajiao3",\
                "xiaolajiaoshouji3", "xiaolajiaosandaishouji"],
                "output_file": "xiaolajiao3_top_comment.xlsx"}
        }


profile_residence_toexcell_task = {
        "xiaolajiao": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "residence",
            "contain": top_comment_toexcel_task["xiaolajiao"]["contain"],
            "output_file": "xiaolajiao_residence.xlsx"},
        "honglajiao": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "residence",
            "contain": top_comment_toexcel_task["honglajiao"]["contain"],
            "output_file": "honglajiao_residence.xlsx"},
        "xiaolajiao3": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "residence",
            "contain": top_comment_toexcel_task["xiaolajiao3"]["contain"],
            "output_file": "xiaolajiao3_residence.xlsx"}
        }


profile_birth_toexcel_task = {
        "xiaolajiao": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "birth",
            "contain": top_comment_toexcel_task["xiaolajiao"]["contain"],
            "output_file": "xiaolajiao_birth.xlsx"},
        "honglajiao": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "residence",
            "contain": top_comment_toexcel_task["honglajiao"]["contain"],
            "output_file": "honglajiao_birth.xlsx"},
        "xiaolajiao3": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "birth",
            "contain": top_comment_toexcel_task["xiaolajiao3"]["contain"],
            "output_file": "xiaolajiao3_birth.xlsx"},

        "nubiya": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "birth",
            "contain": ["nubiya"],
            "output_file": "nubiya_birth.xlsx"},

        "tianyu": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "birth",
            "contain": ["tianyu"],
            "output_file": "tianyu_birth.xlsx"},

        "jiwu": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "birth",
            "contain": ["jiwu"],
            "output_file": "jiwu_birth.xlsx"},

        "thl": {"db": "bbscrawl1",
            "collection": "profile",
            "key": "birth",
            "contain": ["thl"],
            "output_file": "tcl_birth.xlsx"}
        }


def profile_birth_toexcel(kws={}):
    db_name = kws["db"]
    mongo_collection = kws["collection"]
    db = con[db_name]
    contain = kws["contain"]
    wb = Workbook()
    ws = wb.create_sheet(0)
    col_ = 0
    row_ = 1
    result = {}
    for profile in db[mongo_collection].find({"forum_id": {"$in": contain}}):
        birth = profile.get(kws["key"])
        print birth
        if birth:
            if "1" in birth:
                birth = birth.split(u"年")[0]
                print birth
                cur_num = result.get(birth)
                if cur_num:
                    result[birth] = cur_num + 1
                else:
                    result[birth] = 1
    for item in result:
        if len(item) > 4: 
            continue
        col_ = 0
        ws.cell(row=row_, column=col_).value = item
        col_ = col_ + 1
        ws.cell(row=row_, column=col_).value = result[item]
        row_ = row_ + 1
    wb.save(kws["output_file"])


def profile_residence_toexcel(kws={}):
    db_name = kws["db"]
    mongo_collection = kws["collection"]
    contain = kws["contain"]
    db = con[db_name]
    wb = Workbook()
    ws = wb.create_sheet(0)
    col_ = 0
    row_ = 1
    print contain
    result = {}  # type: {"address": num}
    for profile in db[mongo_collection].find({"forum_id": {"$in": contain}}):
        residence = profile.get("residence")
        if residence:
            print residence
            residence = residence.split("-")[0]
            cur_num = result.get(residence)
            if cur_num:
                result[residence] = cur_num + 1
            else:
                result[residence] = 1
    for item in result:
        col_ = 0
        ws.cell(row=row_, column=col_).value = item
        print item
        col_ = col_ + 1
        ws.cell(row=row_, column=col_).value = result[item]
        row_ = row_ + 1
    wb.save(kws["output_file"])


def top_comment_toexcel(kws={}):
    mongo_collection = kws["collection"]
    output_file = kws["output_file"]
    contain = kws["contain"]
    wb = Workbook()
    ws = wb.create_sheet(0)
    db = con["bbscrawl1"]
    count = 0
    col_ = 0
    row_ = 1
    for post in db[mongo_collection].find({"website": "tieba", "forum_id": {"$in": ["oppo", "oppofind5"]}})\
            .sort("comment_num", pymongo.DESCENDING):
        col_ = 0
        if count < 100:
            ws.cell(row=row_, column=col_).value = post["title"]
            col_ = col_ + 1
            ws.cell(row=row_, column=col_).value = post["author"]
            col_ = col_ + 1
            ws.cell(row=row_, column=col_).value = post["content"]
            col_ = col_ + 1
            ws.cell(row=row_, column=col_).value = post["forum_name"]
            col_ = col_ + 1
            ws.cell(row=row_, column=col_).value = post["forum_id"]
            col_ = col_ + 1
            ws.cell(row=row_, column=col_).value = post["url"]
            col_ = col_ + 1
            ws.cell(row=row_, column=col_).value = post.get("comment_num")
            # count = count + 1
            row_ = row_ + 1

    wb.save("oppo.xlsx")


def tieba_comment():
    wb = Workbook()
    ws = wb.create_sheet(0)
    row_ = 1
    # col_ = 0
    f = open("/home/dz/fram_crawl/crawl_proxy/comments.txt")
    count = 0
    for line in f:
        if count > 3000:
            count = count + 1
            continue
        count = count + 1
        col_ = 0
        if "iamsplit" not in line:
            continue
        if "var" in line:
            continue
        line = line.split("iamsplit")[0]  # .strip()
        line = line.encode("ISO-8859-1").decode("GBK").encode("utf-8")
        line = line.strip()
        if len(str(line)) > 0 and "(" not in str(line) and "http" not in str(line) and "www" not in str(line) and "{" not in str(line):
            print line
            print row_
            ws.cell(row=row_, column=0).value = line
            print ws.cell(row=row_, column=1).value
            row_ = row_ + 1
    wb.save("out.xlsx")


def collection_toexcel(kws={}):
    wb = Workbook()
    ws = wb.create_sheet(0)
    mongo_db = mongo_con["bbscrawl1"]
    template = kws["template"]
    find_ = kws["find"]
    row_ = 0
    col_ = 0
    for key in template:
        print template[key]
        ws.cell(row=0, column=col_).value = template[key]
        col_ = col_ + 1
    row_ = row_ + 1
    for item in mongo_db["post"].find(find_):
        col_ = 0
        for key in template:
            ws.cell(row=row_, column=col_).value = item.get(key)
            col_ = col_ + 1
        row_ = row_ + 1
    wb.save("db.xlsx")


if __name__ == "__main__":
    # tieba_comment()
    top_comment_toexcel(top_comment_toexcel_task["xiaolajiao3"])
    # profile_residence_toexcel(profile_residence_toexcell_task["xiaolajiao3"])
    # profile_birth_toexcel(profile_birth_toexcel_task["thl"])
    #mongo_collection = {
    #        "template": {
    #            "_id": u"ID",
    #            "website": u"website",
    #            "uid": u"uid",
    #            "url": u"url",
    #            "comment_num": u"comment_num",
    #            "author": u"author",
    #            "content": u"content",
    #            "forum_name": u"forum_name",
    #            "forum_id": u"type",
    #            "t": u"time"
    #            },
    #        "find": {"forum_id":\
    #                {"$in":\
    #                ["xiaolajiaoshoujiguanwang"]
    #                # ["xiaolajiaom1", "xiaolajiao1s", "xiaolajiaom1y"]
    #                # ["honglajiao"]
    #                # ["xiaolajiao2", "xiaolajiaom2"]
    #                # ["xiaolajiao3", "xiaolajiaoshouji3", "xiaolajiaom3"]
    #                #["xiaolajiao", "beidouxiaolajiao", "beidouxingxiaolajiao", "yuxinxiaolajiao", "xiaolajiaom1", "xiaolajiao1s", "xiaolajiaom1y", "xiaolajiqoq1"]
    #                }}
    #        }
    #collection_toexcel(mongo_collection)
