#coding=utf8

from openpyxl import Workbook
import jieba
import random
import sys
reload(sys)
sys.setdefaultencoding("utf-8")

number = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
letter = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "g", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]


def split_word(word_list):
    dic = {}
    for comment in word_list:
        try:
            comment = comment.encode("ISO-8859-1").decode("GBK").encode("utf-8")
        except Exception, e:
            print e, comment
            continue
        for word in jieba.cut(comment, cut_all=False):
            if len(word) > 1 and word not in stopword_list and word != "iamsplit":
                if word[0] in number or word[0] in letter:
                    continue
                info = dic.get(word, [0])
                info[0] = info[0] + 1
            else:
                continue
            if comment not in info[1:]:
                info.append(comment)
            dic[word] = info
    print "split end"
    l = sorted(dic.items(), key=lambda dic: dic[1][0])
    print "sorted end"
    return list(reversed(l)) # [(word:[num,comment...]), ...]


def word_tail(split_words):
    temp = split_words          # type as:[(word:[num,comment]), ...]
    for item in temp:
        content_num = 0
        len_item = len(item[1][1:]) # 包含该分词的评论个数
        for comment in item[1][1:]:
            content_num = content_num + 1
            is_append = False
            tail_num = 3
            count_ = 1
            for word in jieba.cut(comment, cut_all=False):
                if len(word) > 1 and word not in stopword_list and word != "iamsplit" and word[0] not in number and word[0] not in letter:
                    if not is_append:
                        if word != item[0]:
                            continue
                        else:
                            is_append = True
                            continue
                    if count_ <= tail_num:
                        count_ = count_ + 1
                        item[1].append(word)
                        continue
                    break
        del item[1][1:len_item + 1]
    return temp


def sum_list(source_list):
    result = {}
    for item in source_list:
        num = result.get(item, None)
        if num:
            result[item] = num + 1
        else:
            result[item] = 1
    l = sorted(result.items(), key=lambda result: result[1])
    return list(reversed(l))


def save_xlsx(dict_split):
    # dict_split type:{word:[num,comments...]}
    wb = Workbook()
    ws = wb.create_sheet(0)
    len_dict = len(dict_split)
    col_ = 0
    row_ = 0
    for item in dict_split:
        col_ = 0
        word = item[0]
        print "item[0]", word
        ws.cell(row=row_, column=col_).value = word
        col_ = col_ + 1
        detail = item[1]
        tail(col_, row_, ws, detail, len_dict)
        row_ = row_ + 1
    print "-----------------------------end----------------------"
    wb.save("out.xlsx")


def tail(col_, row_, ws, detail, len_dict):
    col_ = col_
    print detail[0]
    ws.cell(row=row_, column=col_).value = "".join([str((detail[0]/float(len_dict))*100), "%"])
    col_ = col_ + 1
    ws.cell(row=row_, column=col_).value = str(detail[0]) 
    col_ = col_ + 1
    for ite in detail[1]:
        ws.cell(row=row_, column=col_).value = ite[0]
        col_ = col_ + 1
        ws.cell(row=row_, column=col_).value = str(ite[1])
        col_ = col_ + 1
        print "ite", ite[0], ite[1]


def load_stopword():
    f = open("../stopword")
    stopword_list = []
    for word in f:
        stopword_list.append(word)
    return stopword_list

stopword_list = load_stopword()


def split_worker(can_iterator):
    a = split_word(can_iterator)
    print "finish split word"
    b = word_tail(a)
    print "finish tail"
    for item in b:
        print "word num ", item[1][0]
        temp = item[1][1:]
        del item[1][1:]
        tail_num = sum_list(temp)
        item[1].append(tail_num)
    c = save_xlsx(b)


if __name__ == "__main__":
    f = open("../fram_crawl/comments_zhidao_prov2.txt")
    a = split_word(f)
    print "finish split word"
    b = word_tail(a)
    print "finish tail"
    for item in b:
        print "word num ", item[1][0]
        temp = item[1][1:]
        del item[1][1:]
        tail_num = sum_list(temp)
        item[1].append(tail_num)
    c = save_xlsx(b)
